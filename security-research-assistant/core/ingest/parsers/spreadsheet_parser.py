"""Spreadsheet parser for CSV and XLSX files."""

import csv
import io
from pathlib import Path

from core.ingest.parsers.base import BaseParser, PageContent, ParseResult
from core.logging import get_logger

log = get_logger(__name__)


class SpreadsheetParser(BaseParser):
    """Parse CSV and XLSX spreadsheet files into searchable text."""

    def parse(self, filepath: Path) -> ParseResult:
        """Parse a spreadsheet file, converting tabular data to text.

        Args:
            filepath: Path to the spreadsheet file.

        Returns:
            ParseResult with text representation of tabular data.
        """
        ext = filepath.suffix.lower()
        if ext == ".csv" or ext == ".tsv":
            return self._parse_csv(filepath, ext)
        elif ext in (".xlsx", ".xls"):
            return self._parse_xlsx(filepath)
        else:
            return ParseResult(
                text_content="",
                warnings=[f"Unsupported spreadsheet format: {ext}"],
            )

    def _parse_csv(self, filepath: Path, ext: str) -> ParseResult:
        """Parse a CSV/TSV file."""
        warnings: list[str] = []
        delimiter = "\t" if ext == ".tsv" else ","

        try:
            raw = filepath.read_text(encoding="utf-8", errors="replace")
        except Exception as e:
            return ParseResult(text_content="", warnings=[f"Failed to read CSV: {e}"])

        try:
            reader = csv.reader(io.StringIO(raw), delimiter=delimiter)
            rows = list(reader)
        except csv.Error as e:
            warnings.append(f"CSV parse error: {e}")
            return ParseResult(text_content=raw, warnings=warnings)

        if not rows:
            return ParseResult(text_content="", warnings=["Empty CSV file."])

        headers = rows[0]
        text_lines = [" | ".join(headers)]
        text_lines.append("-" * len(text_lines[0]))

        for row in rows[1:]:
            if headers:
                pairs = [f"{h}: {v}" for h, v in zip(headers, row) if v.strip()]
                text_lines.append(" | ".join(pairs))
            else:
                text_lines.append(" | ".join(row))

        text_content = "\n".join(text_lines)
        metadata = {
            "row_count": len(rows) - 1,
            "column_count": len(headers),
            "headers": headers,
            "format": ext.lstrip("."),
        }

        pages = [PageContent(page_number=1, text=text_content, tables=[{
            "headers": headers,
            "rows": rows[1:],
        }])]

        log.info("csv_parsed", filepath=str(filepath), rows=len(rows) - 1)
        return ParseResult(
            text_content=text_content, pages=pages, metadata=metadata, warnings=warnings,
        )

    def _parse_xlsx(self, filepath: Path) -> ParseResult:
        """Parse an XLSX file sheet-by-sheet."""
        warnings: list[str] = []
        pages: list[PageContent] = []
        all_text_parts: list[str] = []

        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
        except Exception as e:
            return ParseResult(text_content="", warnings=[f"Failed to open XLSX: {e}"])

        sheet_names = wb.sheetnames
        metadata: dict = {
            "sheet_names": sheet_names,
            "format": "xlsx",
        }

        total_rows = 0
        for sheet_name in sheet_names:
            try:
                ws = wb[sheet_name]
                rows: list[list[str]] = []
                for row in ws.iter_rows(values_only=True):
                    str_row = [str(c) if c is not None else "" for c in row]
                    if any(c.strip() for c in str_row):
                        rows.append(str_row)

                if not rows:
                    continue

                headers = rows[0]
                text_lines = [f"[Sheet: {sheet_name}]"]
                text_lines.append(" | ".join(headers))
                for row in rows[1:]:
                    pairs = [f"{h}: {v}" for h, v in zip(headers, row) if v.strip()]
                    text_lines.append(" | ".join(pairs))

                sheet_text = "\n".join(text_lines)
                all_text_parts.append(sheet_text)
                pages.append(PageContent(
                    page_number=len(pages) + 1,
                    text=sheet_text,
                    tables=[{"headers": headers, "rows": rows[1:]}],
                ))
                total_rows += len(rows) - 1

            except Exception as e:
                warnings.append(f"Error reading sheet '{sheet_name}': {e}")

        wb.close()

        metadata["row_count"] = total_rows
        text_content = "\n\n".join(all_text_parts)

        log.info("xlsx_parsed", filepath=str(filepath), sheets=len(sheet_names), rows=total_rows)
        return ParseResult(
            text_content=text_content, pages=pages, metadata=metadata, warnings=warnings,
        )
